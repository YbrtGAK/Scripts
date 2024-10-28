# -*- coding: utf-8 -*-
"""
Created on Thu Jan 11 14:18:37 2024

@author: yberton
"""

"""
This script allows the user to pre-fill an excel bibliography file with 
data from the zotero's library.
"""

#imports
from pyzotero import zotero
import pandas as pd
from openpyxl import load_workbook

#Get the zotero library
zot = zotero.Zotero('12936755', 'user', 'cTkuxxmRS1NhgnUlCOG096cB') # library
cols = zot.collections() # Lists of all collections within the library

#Get the items in the "Experiences" collection
items = [item['data'] for item in zot.collection_items("YSTNN8K4") if item['data']['itemType'] == "journalArticle"]

#Get the excel bibliography file in a dataframe
path = "C:/Users/yberton/Documents/Bibliographie/Bibliographie.xlsm"
df = pd.read_excel(path)

#Load the excel bibliography in a openpyxl variable
biblio = load_workbook(path, read_only=False, keep_vba=True)
feuil1 = biblio['Feuil1'] #In particular we load the first sheet

def update(items,df):
    
    # Empty list to collect dataframes
    dfs = []
    #We go through publications withing "Experiences" collection
    for item in items : 
        #If the publication is not in the excel bibliography
        if item['url'] not in list(df['Ref'].values):
            #We add it :
            creators = str([creator['lastName'] for creator in item['creators']]).replace("'","")
            row = {'Ref' : [item['url']], 'Name' : [item['title']], 'Authors' : creators, 'Date' : [item ['date'].split('-')[0]]}
            dfnew = pd.DataFrame.from_dict(row, orient='columns')
            dfs.append(dfnew)
    return(dfs)
dfs = update(items,df)

try :
    df_new = pd.concat(dfs).reset_index(drop=True)
    for c in range(0,len(df_new.columns)) : 
        for index,r in enumerate(df_new.iloc[:,c], start=2) :
            feuil1.cell(row=index+len(df),column=c+1,value=r)
    biblio.save(path)
    print('Succesfully updated :)')
            
except ValueError : print("Already up to date :)")
